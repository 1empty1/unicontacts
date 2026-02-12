import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class DataExporter:
    def __init__(self, db_manager):
        self.db = db_manager

    def _create_default_template(self, filename="template.xlsx"):
        """
        Создает базовый шаблон с заголовком и шапкой, если он отсутствует.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сотрудники"
        

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))


        ws['A1'] = "ВЕДОМОСТЬ СОТРУДНИКОВ УНИВЕРСИТЕТА"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


        ws['A2'] = f"Дата создания отчета: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True, color='808080')
        ws.merge_cells('A2:G2')
        

        headers = ["ID", "ФИО", "Телефон", "Отдел", "Должность", "Корпус", "Кабинет"]
        header_fill = PatternFill(start_color="3B8ED0", end_color="3B8ED0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        wb.save(filename)

    def export_to_excel(self, employees, out_path):
        """
        Экспорт в Excel с использованием шаблона и корректными типами данных.
        """
        if not employees:
            return False,
        try:
            employees = self.db.get_all_employees()
            if not employees:
                return False,

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            if not out_path:
                out_path = f"export_{ts}.xlsx"


            template_name = "template.xlsx"
            if not os.path.exists(template_name):
                self._create_default_template(template_name)


            wb = openpyxl.load_workbook(template_name)
            ws = wb.active
            

            ws['A2'] = f"Дата создания отчета: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
            
            start_row = 4 
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))


            center_cols = [0, 5, 6] 

            for i, emp in enumerate(employees):
                row_num = start_row + i
                

                processed_emp = list(emp)
                

                try: processed_emp[0] = int(emp[0])
                except: pass

                try:
                    if str(emp[5]).isdigit():
                        processed_emp[5] = int(emp[5])
                except:
                    pass

                try:
                    if str(emp[6]).isdigit():
                        processed_emp[6] = int(emp[6])
                except:
                    pass

                for col_index, value in enumerate(processed_emp):
                    col_num = col_index + 1 
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    cell.value = value
                    cell.border = thin_border
                    
                    if col_index in center_cols:
                        cell.alignment = Alignment(horizontal='center', vertical='top')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='top')

            for i, column_cells in enumerate(ws.columns, 1):
                length = 0
                for cell in column_cells:
                    try:
                        if cell.value:
                            curr_len = len(str(cell.value))
                            if curr_len > length:
                                length = curr_len
                    except:
                        pass
                
                col_letter = get_column_letter(i)
                if i in [1, 6, 7]: 
                    width = min(length + 2, 12)
                elif i in [2, 4]: 
                    width = min(length + 2, 40)
                else:
                    width = length + 2
                
                ws.column_dimensions[col_letter].width = width

            wb.save(out_path)
            return True, f"Данные успешно экспортированы в {os.path.abspath(out_path)}"
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, f"Ошибка экспорта: {e}"